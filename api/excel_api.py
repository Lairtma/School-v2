import os
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook


def create_empty_excel(columns: list, filename: str, sheet_name: str = 'Лист1'):
    df = pd.DataFrame(columns=columns)

    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', filename)
    excel_writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 0))
    excel_writer._save()

    return filepath

def parse_excel_to_dict_list(filepath: str, sheet_name='Лист1'):
    # Загружаем Excel файл в DataFrame
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Преобразуем DataFrame в список словарей
    dict_list = df.to_dict(orient='records')

    return dict_list

def get_data_to_exel(filepath: str):
    info = parse_excel_to_dict_list(filepath)
    classes = info[4]
    monday = info[5:14]
    tuesday = info[14:23]
    wednesday = info[23:32]
    thursday = info[32:41]
    friday = info[41:50]
    using_columns = [f"Unnamed: {x}" for x in range(2, 18)] + [f"Unnamed: {x}" for x in range(20, 46)]
    classes = [classes[x] for x in using_columns]
    for i in range(9):
        monday[i] = [monday[i][x] for x in using_columns]
        tuesday[i] = [tuesday[i][x] for x in using_columns]
        wednesday[i] = [wednesday[i][x] for x in using_columns]
        thursday[i] = [thursday[i][x] for x in using_columns]
        friday[i] = [friday[i][x] for x in using_columns]
    schedule = {"Понедельник": dict(), "Вторник": dict(), "Среда": dict(), "Четверг": dict(), "Пятница": dict()}
    for i in range(len(classes)):
        schedule["Понедельник"][classes[i]] = dict()
        for j in range(9):
            if type(monday[j][i]) == str:
                schedule["Понедельник"][classes[i]][j + 1] = monday[j][i]
        schedule["Вторник"][classes[i]] = dict()
        for j in range(9):
            if type(tuesday[j][i]) == str:
                schedule["Вторник"][classes[i]][j + 1] = tuesday[j][i]
        schedule["Среда"][classes[i]] = dict()
        for j in range(9):
            if type(wednesday[j][i]) == str:
                schedule["Среда"][classes[i]][j + 1] = wednesday[j][i]
        schedule["Четверг"][classes[i]] = dict()
        for j in range(9):
            if type(thursday[j][i]) == str:
                schedule["Четверг"][classes[i]][j + 1] = thursday[j][i]
        schedule["Пятница"][classes[i]] = dict()
        for j in range(9):
            if type(friday[j][i]) == str:
                schedule["Пятница"][classes[i]][j + 1] = friday[j][i]
    return schedule


def create_excel_from_dict_list(dict_list: list, output_filename: str, sheet_name='Лист1'):
    # Создаем директорию, если она не существует
    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    # Настраиваем стили для красивого вида
    header_style = NamedStyle(name='header')
    header_style.font = Font(bold=True, color='FFFFFF')
    header_style.alignment = Alignment(horizontal='center', vertical='center')
    header_style.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border_style = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    header_style.border = border_style

    cell_style = NamedStyle(name='cell')
    cell_style.alignment = Alignment(horizontal='left', vertical='center')
    cell_style.border = border_style

    for cell in ws[1]:  # Применяем стиль к заголовкам
        cell.style = header_style

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.style = cell_style

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath


if __name__ == '__main__':
    get_data_to_exel("test.xlsx")